"""
report.py — render scoring results as a human-readable Markdown scorecard and
an Excel shortlist. The scorecard always explains *why* (evidence per sub-criterion).
Reads category/sub-criterion labels straight from the result, so it stays in sync
with model.yaml automatically.
"""

from __future__ import annotations
from pathlib import Path


def scorecard_markdown(result: dict, role: dict) -> str:
    L = []
    wow = " 🏆 **WOW**" if result["wow"] else ""
    L.append(f"# Scorecard — {result['candidate']}{wow}")
    L.append("")
    L.append(f"**Role:** {role.get('title')} (`{role.get('id')}`) — {role.get('client','')}")
    L.append(f"**Total score:** **{result['total_score']}/100**  ·  "
             f"**Status:** {result['status']}  ·  "
             f"**Data completeness:** {int(result['data_completeness']*100)}%")
    L.append("")

    np = result.get("no_poach", {})
    if np.get("triggered"):
        L.append(f"> ⛔ **NO-POACH** — currently at active client: **{', '.join(np['companies'])}**. "
                 f"Do not pursue without sign-off (don't recruit away from our own clients).")
        L.append("")

    if not result["gate_passed"]:
        L.append(f"> ⚠️ **Missing must-have(s):** {', '.join(result['missing_must_haves'])}")
        L.append("")

    feas = result.get("feasibility", [])
    if feas:
        icon = {"pass": "✅", "review": "⚠️", "unknown": "❔"}
        L.append("## Feasibility gates _(pass/fail filters — not scored)_")
        L.append("")
        for f in feas:
            L.append(f"- {icon.get(f['status'],'•')} **{f['label']}** — {f['status']}: {f['evidence']}")
        L.append("")

    L.append("## Category breakdown")
    L.append("")
    L.append("| Category | Weight | Score |")
    L.append("|---|---:|---:|")
    for cat in result["categories"].values():
        L.append(f"| {cat['label']} | {cat['weight']:.2f} | {cat['score']:.0f} |")
    L.append("")

    L.append("## Evidence")
    for cat in result["categories"].values():
        L.append("")
        L.append(f"### {cat['label']} — {cat['score']:.0f}/100  _(weight {cat['weight']:.2f})_")
        for sub in cat["subcriteria"].values():
            tag = "" if sub["provenance"] == "explicit" else " _(no data → neutral)_"
            flag = f"  \n  ⚠️ _{sub['fairness_flag']}_" if sub.get("fairness_flag") else ""
            L.append(f"- **{sub['label']}** _(w {sub['weight']:.2f})_ — "
                     f"{sub['score']:.0f}: {sub['evidence']}{tag}{flag}")

    L.append("")
    L.append("---")
    L.append("_Scores are decision support, not a decision. Inferred/neutral fields are flagged; "
             "verify before outreach._")
    return "\n".join(L)


def write_markdown(result: dict, role: dict, out_dir: str | Path) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    safe = "".join(ch if ch.isalnum() else "_" for ch in result["candidate"]).strip("_")
    path = out_dir / f"scorecard_{safe}.md"
    path.write_text(scorecard_markdown(result, role), encoding="utf-8")
    return path


def write_shortlist_xlsx(results: list[dict], role: dict, out_dir: str | Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Shortlist"

    headers = ["Rank", "Candidate", "Total", "Status", "Wow",
               "1. Profile", "2. Skills", "3. Credibility", "4. Availability", "5. Mobility",
               "Missing must-haves", "Completeness"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    cat_order = ["professional_profile", "skills_qualifications", "credibility_distinction",
                 "availability_retention", "mobility"]
    ranked = sorted(results, key=lambda r: r["total_score"], reverse=True)
    for i, r in enumerate(ranked, start=1):
        cs = r["category_scores"]
        row = [i, r["candidate"], r["total_score"], r["status"], "🏆" if r["wow"] else ""]
        row += [round(cs[k]) for k in cat_order]
        row += [", ".join(r["missing_must_haves"]), f"{int(r['data_completeness']*100)}%"]
        for col, val in enumerate(row, start=1):
            ws.cell(row=i + 1, column=col, value=val)

    widths = [6, 22, 8, 26, 6, 10, 10, 12, 12, 10, 30, 12]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    path = out_dir / f"shortlist_{role.get('id','role')}.xlsx"
    wb.save(path)
    return path
