"""
make_tuning_sheet.py — generate an editable Excel weight-tuning workbook from model.yaml.

You edit the yellow cells (category weights + sub-weights); the sheet live-checks
that each block sums to 1.00 and auto-computes each criterion's effective points.
Send it back and `sync_weights.py` writes your numbers into model.yaml.

    python make_tuning_sheet.py            # -> tuning/Scoring_Model_Tuning.xlsx
"""

from __future__ import annotations
from pathlib import Path
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

HERE = Path(__file__).parent
CAT_COLORS = ["1F4E79", "2E75B6", "4472C4", "2F5597", "1F6E51"]

# styles
WHITE_BOLD = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BOLD = Font(name="Arial", bold=True, size=10)
BODY = Font(name="Arial", size=10)
ITAL = Font(name="Arial", italic=True, size=9, color="595959")
EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")    # yellow = editable
CALC_FILL = PatternFill("solid", fgColor="EDEDED")     # grey = computed
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def build():
    model = yaml.safe_load((HERE / "model.yaml").read_text(encoding="utf-8"))
    cats = model["categories"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Weight Tuning"

    # ── Title + instructions ──────────────────────────────────────────────────
    ws["A1"] = "SOURCING — Scoring Model Tuning Sheet"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A2"] = ("Edit the YELLOW cells only. Category weights (top) must sum to 1.00. "
                "Each category's sub-weights must sum to 1.00. Grey cells auto-calculate.")
    ws["A2"].font = ITAL
    ws["A3"] = ("'Effective pts' = category weight × sub-weight × 100 (the criterion's real "
                "share of the /100 score). ✅ = scored now from CV · ⏳ = awaits LinkedIn/Zoho data.")
    ws["A3"].font = ITAL

    dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=1,
                        allow_blank=False,
                        error="Weights must be between 0 and 1.", errorTitle="Invalid weight")
    ws.add_data_validation(dv)

    # ── Section A: category weights ───────────────────────────────────────────
    r = 5
    ws.cell(r, 1, "CATEGORY WEIGHTS").font = WHITE_BOLD
    for c in range(1, 5):
        ws.cell(r, c).fill = HDR_FILL
    r += 1
    for col, h in enumerate(["Category", "Weight (edit)", "Check"], start=1):
        cell = ws.cell(r, col, h)
        cell.font = BOLD
        cell.fill = PatternFill("solid", fgColor="D6E4F0")
        cell.border = BORDER
    cat_weight_row = {}      # cat_id -> row holding its editable weight (col B)
    first_cat_row = r + 1
    for i, (cat_id, cat) in enumerate(cats.items()):
        r += 1
        ws.cell(r, 1, cat["label"]).font = BODY
        wc = ws.cell(r, 2, round(cat["weight"], 4))
        wc.fill = EDIT_FILL; wc.font = BODY; wc.border = BORDER
        wc.number_format = "0.00"
        dv.add(wc)
        cat_weight_row[cat_id] = r
    last_cat_row = r
    r += 1
    ws.cell(r, 1, "TOTAL").font = BOLD
    tot = ws.cell(r, 2, f"=SUM(B{first_cat_row}:B{last_cat_row})")
    tot.font = BOLD; tot.fill = CALC_FILL; tot.number_format = "0.00"; tot.border = BORDER
    chk = ws.cell(r, 3, f'=IF(ABS(B{r}-1)<0.001,"✓ OK","✗ must = 1.00")')
    chk.font = BOLD
    cat_total_check_row = r

    # ── Section B: sub-criteria weights per category ──────────────────────────
    r += 3
    ws.cell(r, 1, "SUB-CRITERIA WEIGHTS").font = WHITE_BOLD
    for c in range(1, 7):
        ws.cell(r, c).fill = HDR_FILL
    r += 1
    headers = ["#", "Sub-criterion", "Sub-weight (edit)", "Effective pts", "Now?", "Data needed (for ⏳)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(r, col, h)
        cell.font = BOLD; cell.fill = PatternFill("solid", fgColor="D6E4F0")
        cell.alignment = CENTER; cell.border = BORDER
    header_row = r

    subtotal_d_rows = []
    for i, (cat_id, cat) in enumerate(cats.items()):
        crow = cat_weight_row[cat_id]
        r += 1
        # block header
        bh = ws.cell(r, 2, f"{cat['label']}   (category weight in B{crow})")
        bh.font = WHITE_BOLD
        for c in range(1, 7):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=CAT_COLORS[i])
        ws.cell(r, 1, f"{i+1}").font = WHITE_BOLD
        block_first = r + 1
        for j, sub in enumerate(cat["subcriteria"], start=1):
            r += 1
            ws.cell(r, 1, j).font = BODY
            lab = ws.cell(r, 2, sub["label"]); lab.font = BODY; lab.alignment = WRAP; lab.border = BORDER
            wc = ws.cell(r, 3, round(sub["weight"], 4))
            wc.fill = EDIT_FILL; wc.font = BODY; wc.number_format = "0.00"; wc.border = BORDER
            dv.add(wc)
            eff = ws.cell(r, 4, f"=$B${crow}*C{r}*100")
            eff.fill = CALC_FILL; eff.font = BODY; eff.number_format = "0.0"; eff.border = BORDER
            scorable = sub.get("scorer") is not None
            st = ws.cell(r, 5, "✅ now" if scorable else "⏳ later")
            st.alignment = CENTER; st.border = BORDER
            nd = ws.cell(r, 6, "" if scorable else sub.get("needs", "richer data"))
            nd.font = ITAL; nd.alignment = WRAP; nd.border = BORDER
        block_last = r
        # subtotal
        r += 1
        ws.cell(r, 2, "Sub-weights total").font = BOLD
        sc = ws.cell(r, 3, f"=SUM(C{block_first}:C{block_last})")
        sc.font = BOLD; sc.fill = CALC_FILL; sc.number_format = "0.00"; sc.border = BORDER
        sd = ws.cell(r, 4, f"=SUM(D{block_first}:D{block_last})")
        sd.font = BOLD; sd.fill = CALC_FILL; sd.number_format = "0.0"; sd.border = BORDER
        subtotal_d_rows.append(r)
        ck = ws.cell(r, 5, f'=IF(ABS(C{r}-1)<0.001,"✓","✗ fix=1.00")')
        ck.font = BOLD; ck.alignment = CENTER
        r += 1  # blank spacer

    # grand total of effective points (should be 100)
    r += 1
    ws.cell(r, 2, "GRAND TOTAL — effective pts").font = BOLD
    grand = ws.cell(r, 4, "=" + "+".join(f"D{x}" for x in subtotal_d_rows))
    grand.font = BOLD; grand.fill = CALC_FILL; grand.number_format = "0.0"; grand.border = BORDER
    gck = ws.cell(r, 5, f'=IF(ABS(D{r}-100)<0.1,"✓ =100","✗ check")')
    gck.font = BOLD; gck.alignment = CENTER

    # ── conditional formatting: red ✗ checks, green ✓ ─────────────────────────
    chk_range = f"C{cat_total_check_row} E{header_row}:E{r}"
    for rng in (f"C{cat_total_check_row}", f"E{header_row+1}:E{r}"):
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'ISNUMBER(SEARCH("✗",{rng.split(":")[0]}))'],
            fill=PatternFill("solid", fgColor="F8CBAD")))
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=[f'ISNUMBER(SEARCH("✓",{rng.split(":")[0]}))'],
            fill=PatternFill("solid", fgColor="C6EFCE")))

    # widths
    for col, w in zip("ABCDEF", [5, 58, 16, 13, 10, 40]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    out = HERE / "tuning" / "Scoring_Model_Tuning.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print("Saved:", out)
    return out


if __name__ == "__main__":
    build()
