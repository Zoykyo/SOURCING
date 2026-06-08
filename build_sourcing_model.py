from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── SHEET 1: Data Sources & Multi-Criteria Model ─────────────────────────────
ws1 = wb.active
ws1.title = "Data Sources & Criteria"

header_font = Font(name="Arial", bold=True, size=11)
body_font   = Font(name="Arial", size=10)
wrap        = Alignment(wrap_text=True, vertical="top")
thin        = Side(style="thin", color="BFBFBF")
border      = Border(bottom=thin)

# Headers
ws1["B1"] = "Data Sources"
ws1["C1"] = "Multi-Criteria Model"
for cell in [ws1["B1"], ws1["C1"]]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")

data_sources = [
    "Potential client list",
    "Resume examples",
    "Job descriptions – Roles",
    "Expertise profile (summary)",
    "Public data – Candidate pool",
    "Compensation survey – Genium",
    "Industry data (surveys)",
    "Past requisition list – 5 years",
    "Mutual referral model",
    "Exclusion list or model",
    None, None, None, None, None, None, None, None, None, None, None, None,
]

criteria = [
    "Career path analysis",
    "Key competencies (must-have vs. nice-to-have) and skills",
    "Employers – Key OEM",
    "Relevant roles and experiences",
    "Region – Past and current studies and work",
    "Education, continuing education, certification and credentials",
    "Seniority level",
    "Behavioral/habit model",
    "Cross-functional or similar domain/role",
    "Bonus model, added value",
    "Career progression analysis (lateral, hierarchical, interests, etc.) vs. what we can offer",
    "Contact analysis (personal vs. manager)",
    "Recent activities",
    "Tenure windows (current and avg. past) and open to work",
    "Probabilities to stay at Kollabtek",
    "Language proficiency analysis",
    "Interest in Kollabtek, clients and sector companies (subscribed, connected, applied in the past, etc.)",
    "Location (studied, worked, travel frequency) – interest in relocating, traveling, and proximity to current role",
    "References, feedback",
    "Awards, distinctions, scholarships, competitions, honors, recognitions, alumni",
    "Volunteering, involvement, causes, community contribution",
    "Publications and patents",
]

for i, (src, crit) in enumerate(zip(data_sources, criteria), start=2):
    row_fill = PatternFill("solid", fgColor="EBF3FB") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    if src:
        ws1.cell(row=i, column=2, value=src).font = body_font
        ws1.cell(row=i, column=2).alignment = wrap
        ws1.cell(row=i, column=2).fill = row_fill
    ws1.cell(row=i, column=3, value=crit).font = body_font
    ws1.cell(row=i, column=3).alignment = wrap
    ws1.cell(row=i, column=3).fill = row_fill

ws1.column_dimensions["B"].width = 38
ws1.column_dimensions["C"].width = 65
ws1.row_dimensions[1].height = 30
for r in range(2, 24):
    ws1.row_dimensions[r].height = 30
ws1.freeze_panes = "B2"

# ── SHEET 2: Multi-Criteria Model ────────────────────────────────────────────
ws2 = wb.create_sheet("Multi-Criteria Model")

cat_headers = [
    "1. Professional Profile\n& Career Trajectory",
    "2. Skills, Expertise\n& Qualifications",
    "3. Credibility, Reputation\n& Distinction",
    "4. Availability\n& Retention Potential",
    "5. Mobility",
]
cat_colors = ["1F4E79", "2E75B6", "4472C4", "2F5597", "1F4E79"]
focus_texts = [
    "Focus: Career path, progression, experience relevance, professional history.",
    "Focus: Capability, technical fit, credentials, match with the opportunity and company.",
    "Focus: Distinction, excellence, credibility, reputation.",
    "Focus: Likelihood of moving, staying, and being reachable.",
    "Focus: Interest, possibility to relocate and travel.",
]

cat_criteria = {
    1: [
        "Employers – Key OEM: List and relevance of companies where they worked",
        "Relevant roles and experiences; current role title vs. what we're proposing",
        "Seniority level vs. what's being sought",
        "Professional career analysis: progression (lateral, hierarchical, interests, etc.) vs. what we can offer",
        "Sub-domain match",
    ],
    2: [
        "Key competencies and skills (must-have vs. nice-to-have)",
        "Key knowledge and standards (SAE, ISO)",
        "Education, continuing education (courses)",
        "Language proficiency analysis",
        "Professional associations, certifications and credentials",
    ],
    3: [
        "Awards, distinctions, scholarships, honors, awards, recognitions, alumni",
        "References, feedback",
        "Demonstrative expertise or domain authority: Speaker, panel, trainer, committees",
        "Competition: student club",
        "Publications and patents",
    ],
    4: [
        "Interest in Kollabtek, clients and sector companies (subscribed, connected, applied in the past, etc.)",
        "Open to work or probability of being open to work",
        "Contact analysis (recruiter, Kollabtek team)",
        "Probabilities to stay at Kollabtek",
        "Tenure windows (current and avg. past)",
    ],
    5: [
        "Interest in relocation",
        "Willingness to travel",
        "Location: proximity to current role",
        "Region – past and current studies and work",
        "Work mode: on-site, hybrid, remote",
    ],
}

# Row 1: Category headers
for col, (header, color) in enumerate(zip(cat_headers, cat_colors), start=2):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

# Row 2: Focus descriptions
focus_fill = PatternFill("solid", fgColor="D6E4F0")
for col, focus in enumerate(focus_texts, start=2):
    cell = ws2.cell(row=2, column=col, value=focus)
    cell.font = Font(name="Arial", italic=True, size=9, color="1F4E79")
    cell.fill = focus_fill
    cell.alignment = Alignment(wrap_text=True, vertical="top")

# Rows 3-7: Criteria
for row_offset, crit_idx in enumerate(range(1, 6), start=3):
    row_fill = PatternFill("solid", fgColor="EBF3FB") if row_offset % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")
    for col in range(1, 6):
        items = cat_criteria[col]
        val = items[crit_idx - 1] if crit_idx - 1 < len(items) else ""
        cell = ws2.cell(row=row_offset, column=col + 1, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.fill = row_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# Rows 9-10: Lateral/bonus considerations (green highlight, span all cols)
green_fill = PatternFill("solid", fgColor="92D050")
lateral_items = [
    "Lateral consideration: Bonus/added value model (A + B + C = Wow!)",
    "Profile enrichment model (the unspoken) – Profiles always contain partial, missing, or incomplete information…",
]
for i, text in enumerate(lateral_items, start=9):
    cell = ws2.cell(row=i, column=2, value=text)
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.fill = green_fill
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

# Column widths & row heights
for col in range(2, 7):
    ws2.column_dimensions[get_column_letter(col)].width = 32
ws2.row_dimensions[1].height = 45
ws2.row_dimensions[2].height = 40
for r in range(3, 11):
    ws2.row_dimensions[r].height = 50
ws2.freeze_panes = "B3"

out = "/sessions/adoring-hopeful-rubin/mnt/outputs/Sourcing model (EN).xlsx"
wb.save(out)
print("Saved:", out)
